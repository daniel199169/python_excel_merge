from openpyxl import Workbook, load_workbook
import time
import json
from datetime import datetime

x='{"INAMI":1,"BELFIRST":2,"A-CGDIM":3,"WUD":4,"":5}'

def merge(a,b):
    provenance=json.loads(x)
    if provenance[a[0].upper()]>provenance[b[0].upper()] :
        temp=a.copy()
        a=b.copy()
        b=temp.copy()

    for i in range(1,35):
        if (i in (9,10,11)) and b[i]!='':  #email
            _array=[9,10,11]
            for index in _array:
                if a[index]==b[i]:
                    break
                elif a[index]=='':
                    a[index]=b[i]
                    break
        elif (i in (14,15)) and b[i]!='':#speciality
            _array=[14,15]
            for index in _array:
                if a[index].upper()==b[i].upper():
                    break
                elif a[index]=='':
                    a[index]=b[i]
                    break
        if (i in (16,17)) and b[i]!='':#phone
            _array=[16,17]
            for index in _array:
                if a[index].upper()==b[i].upper():
                    break
                elif a[index]=='':
                    a[index]=b[i]
                    break
        elif (i in (21,27,32,33,34)) and b[i]!='':#address
            _array=[21,27,32,33,34]
            for index in _array:
                if a[index].upper()==b[i].upper():
                    break
                elif a[index]=='':
                    a[index]=b[i]
                    break
        elif (i in (24,30)) and b[i]!='':#zip
            _array=[24,30]
            for index in _array:
                if a[index]==b[i]:
                    break
                elif a[index]=='':
                    a[index]=b[i]
                    break
        elif (i in (25,31)) and b[i]!='':#city
            _array=[25,31]
            for index in _array:
                if a[index].upper()==b[i].upper():
                    break
                elif a[index]=='':
                    a[index]=b[i]
                    break
        elif (i in (22,28)) and b[i]!='':#Numéro
            _array=[22,28]
            for index in _array:
                if a[index]==b[i]:
                    break
                elif a[index]=='':
                    a[index]=b[i]
                    break
        elif (i in (23,29)) and b[i]!='':#Boite
            _array=[23,29]
            for index in _array:
                if a[index]==b[i]:
                    break
                elif a[index]=='':
                    a[index]=b[i]
                    break
        elif a[i]=='' and b[i]!='':
            a[i]=b[i]
    return a


read_file = "MEDICAL_TOTAL-c.xlsx"
# read_file = "clean-file-test.xlsx"
write_file="arrangement1.xlsx"


start_time=time.time()
read_xlsx=load_workbook(read_file, read_only=True)      #creating a workbook from file

print("read time : ",time.time()-start_time)
read_sheet=read_xlsx['Full']
result_array=[]
i=0
for row in read_sheet:
    i+=1
    j=0
    one_row=[]
    for cell in row:
        j+=1
        if j>35:break
        _str=cell.value
        if i>1 and j==13 and _str!=None and _str!='':
            # _date = cell.value.date()
            # _str = str(_date.strftime("%d/%m/%Y"))
            try:
                string=str(_str)
                _y=string[0:4]
                _m=string[5:7]
                _d=string[8:10]
                _str=str(int(_d))+'/'+str(int(_m))+'/'+_y
            except:
                _str=cell.value
                pass
        if _str==None or _str=='':
            _str=''
        one_row.append(str(_str))
    if i==1:
        result_array.append(one_row)
    else:
        index=1
        equal=False
        for elementofarray in result_array:
            if str(one_row[13])==str(elementofarray[13]) :
                equal=True
                break
            elif one_row[5].upper() == elementofarray[5].upper() and one_row[6].upper()==elementofarray[6].upper() and one_row[24]==elementofarray[24]: #postal code
                equal=True
                break
            elif one_row[5].upper() == elementofarray[5].upper() and one_row[6].upper()==elementofarray[6].upper() and one_row[9]==elementofarray[9]: #email
                equal=True
                break
            elif one_row[5].upper() == elementofarray[5].upper() and one_row[6].upper()==elementofarray[6].upper() and one_row[21]==elementofarray[21]: #address1
                equal=True
                break
            elif one_row[5].upper() == elementofarray[5].upper() and one_row[6].upper()==elementofarray[6].upper() and one_row[16]==elementofarray[16]: #telephone
                equal=True
                break
            elif one_row[5].upper() == elementofarray[5].upper() and one_row[6].upper()==elementofarray[6].upper():
                equal=True
                break
            index += 1
        if equal==True :
            merge_array = merge(one_row,result_array[index-1])
            result_array.insert(index-1,merge_array)
            result_array.pop(index)
        else :
            result_array.append(one_row)
        print("current row : ",str(i))
    # if i>500:break

write_xlsx=Workbook()     #creating an empty workbook
write_sheet = write_xlsx.active
write_sheet.title = "result"
for index in range(1,len(result_array)+1):
    for col in range(1,36):
        write_sheet.cell(row=index,column=col).value = result_array[index-1][col-1]

write_xlsx.save(write_file)

print("all progress time : ",time.time()-start_time)